
#!/usr/bin/env python3
"""
PDF Vector Importer QA runner

Purpose:
- Read the v3 QA workbook
- Validate workbook schema
- Select tests by platform / priority / automation
- Run AUTO/HYBRID tests through external adapters when configured
- Emit CSV + JSON results
- Optionally write a copy of the workbook with updated Status/Notes

Honest limitation:
This harness does not magically drive SketchUp or FreeCAD by itself.
Actual host execution must be provided through adapter commands or wrapper scripts.
"""

import argparse
import csv
import json
import shlex
import subprocess
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

REQUIRED_HEADERS = [
    "Test ID","Requirement Ref","Category","Priority","Platform","Automation",
    "Test Name","Input","Steps","Expected Result","Metrics / Tolerance","Status","Notes"
]

DEFAULT_SHEETS = ["SketchUp Tests", "FreeCAD Tests", "Blender Tests", "LibreCAD Tests"]

PLATFORM_TO_ADAPTER_KEY = {
    "SU": "sketchup",
    "FC": "freecad",
    "BL": "blender",
    "LC": "librecad",
}

DEFAULT_TEST_ROWS = {
    "SketchUp Tests": {
        "test_id": "SU-SMOKE-001",
        "platform": "SU",
        "name": "SketchUp smoke test placeholder",
    },
    "FreeCAD Tests": {
        "test_id": "FC-SMOKE-001",
        "platform": "FC",
        "name": "FreeCAD smoke test placeholder",
    },
    "Blender Tests": {
        "test_id": "BL-SMOKE-001",
        "platform": "BL",
        "name": "Blender smoke test placeholder",
    },
    "LibreCAD Tests": {
        "test_id": "LC-SMOKE-001",
        "platform": "LC",
        "name": "LibreCAD smoke test placeholder",
    },
}

@dataclass
class TestCase:
    test_id: str
    requirement_ref: str
    category: str
    priority: str
    platform: str
    automation: str
    test_name: str
    input_file: str
    steps: str
    expected_result: str
    metrics: str
    status: str = ""
    notes: str = ""
    sheet_name: str = ""
    row_number: int = 0

@dataclass
class TestResult:
    test_id: str
    platform: str
    priority: str
    automation: str
    result: str
    runtime_seconds: float
    notes: str

def normalize(s):
    return "" if s is None else str(s).strip()


def init_workbook(path: str):
    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in [*DEFAULT_SHEETS, "All Tests"]:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(REQUIRED_HEADERS)

    for sheet_name in DEFAULT_SHEETS:
        row = DEFAULT_TEST_ROWS[sheet_name]
        values = [
            row["test_id"],
            "REQ-SMOKE-001",
            "Smoke",
            "P0",
            row["platform"],
            "AUTO",
            row["name"],
            "path/to/test.pdf",
            "Run importer using default preset",
            "Importer completes without fatal errors",
            "No crash; geometry count >= 1",
            "",
            "",
        ]
        wb[sheet_name].append(values)
        wb["All Tests"].append(values)

    wb.save(workbook_path)

def load_config(path: Optional[str]) -> Dict:
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as f:
        if path.lower().endswith(".json"):
            return json.load(f)
        raise ValueError("Only JSON config is supported in this starter harness.")

def validate_headers(ws):
    actual = [normalize(ws.cell(1, c).value) for c in range(1, len(REQUIRED_HEADERS)+1)]
    if actual != REQUIRED_HEADERS:
        raise ValueError(f"{ws.title}: header mismatch.\nExpected: {REQUIRED_HEADERS}\nActual:   {actual}")

def iter_tests(workbook_path: str, sheets: List[str]) -> List[TestCase]:
    wb = load_workbook(workbook_path)
    tests: List[TestCase] = []
    missing_sheets: List[str] = []
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            missing_sheets.append(sheet_name)
            continue
        ws = wb[sheet_name]
        validate_headers(ws)
        for r in range(2, ws.max_row + 1):
            row = [normalize(ws.cell(r, c).value) for c in range(1, len(REQUIRED_HEADERS)+1)]
            if not row[0]:
                continue
            tests.append(TestCase(
                test_id=row[0],
                requirement_ref=row[1],
                category=row[2],
                priority=row[3],
                platform=row[4],
                automation=row[5],
                test_name=row[6],
                input_file=row[7],
                steps=row[8],
                expected_result=row[9],
                metrics=row[10],
                status=row[11],
                notes=row[12],
                sheet_name=sheet_name,
                row_number=r,
            ))
    if missing_sheets:
        print(
            "Warning: workbook missing optional sheet(s): "
            + ", ".join(missing_sheets),
            file=sys.stderr,
        )
    if not tests:
        raise ValueError(
            "No test rows found. Ensure at least one platform test sheet exists "
            f"with required headers: {', '.join(sheets)}"
        )
    return tests

def select_tests(tests: List[TestCase], platform: Optional[str], priority: Optional[str], automation: Optional[str]) -> List[TestCase]:
    out = []
    for t in tests:
        if platform and t.platform != platform:
            continue
        if priority and t.priority != priority:
            continue
        if automation and t.automation != automation:
            continue
        out.append(t)
    return out

def build_command(template: str, test: TestCase, config: Dict) -> str:
    replacements = {
        "test_id": test.test_id,
        "platform": test.platform,
        "priority": test.priority,
        "automation": test.automation,
        "category": test.category,
        "test_name": test.test_name,
        "input": test.input_file,
        "steps": test.steps,
        "expected": test.expected_result,
        "metrics": test.metrics,
        "workspace": str(Path(config.get("workspace", ".")).resolve()),
    }
    return template.format(**replacements)

def parse_adapter_json(stdout_text: str) -> Optional[Dict]:
    s = (stdout_text or "").strip()
    if not s:
        return None
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError:
        pass

    first = s.find("{")
    last = s.rfind("}")
    if first >= 0 and last > first:
        candidate = s[first:last + 1]
        try:
            obj = json.loads(candidate)
            return obj if isinstance(obj, dict) else None
        except json.JSONDecodeError:
            return None
    return None

def classify_adapter_result(returncode: int, payload: Optional[Dict]) -> Tuple[str, str]:
    if returncode != 0:
        return "Fail", f"Adapter process exit code {returncode}."

    if not payload:
        return "Pass", "Adapter returned exit code 0."

    status = normalize(payload.get("status")).upper()
    message = normalize(payload.get("message"))

    if status in ("PASS", "OK", "SUCCESS"):
        return "Pass", message or f"Adapter status={status}."
    if status in ("DRY_RUN", "BLOCKED", "MANUAL"):
        return "Blocked", message or f"Adapter status={status}."
    if status in ("FAIL", "FAILED", "ERROR", "TIMEOUT", "UNKNOWN"):
        return "Fail", message or f"Adapter status={status}."
    if status:
        return "Fail", message or f"Unrecognized adapter status={status}."
    return "Pass", message or "Adapter returned exit code 0."

def run_one(test: TestCase, config: Dict, dry_run: bool) -> TestResult:
    start = time.perf_counter()

    if test.automation == "MANUAL":
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Blocked",
            runtime_seconds=0.0,
            notes="Manual test. Not executed by harness."
        )

    adapters = config.get("adapters", {})
    platform_key = PLATFORM_TO_ADAPTER_KEY.get(test.platform)
    if not platform_key:
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Blocked",
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes=f"Unsupported platform code: {test.platform}",
        )
    adapter = adapters.get(platform_key)

    if dry_run:
        note = "Dry run only."
        if adapter and adapter.get("command"):
            note += f" Adapter command template found for {platform_key}."
        else:
            note += f" No adapter configured for {platform_key}."
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Blocked",
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes=note,
        )

    if not adapter or not adapter.get("command"):
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Blocked",
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes=f"No adapter configured for {platform_key}.",
        )

    cmd = build_command(adapter["command"], test, config)
    try:
        completed = subprocess.run(
            shlex.split(cmd),
            cwd=config.get("workspace", None),
            capture_output=True,
            text=True,
            timeout=int(adapter.get("timeout_seconds", 900)),
            check=False,
        )
        stdout = (completed.stdout or "").strip()
        stderr = (completed.stderr or "").strip()
        payload = parse_adapter_json(stdout)
        result, note = classify_adapter_result(completed.returncode, payload)

        note_parts = []
        if note:
            note_parts.append(note)
        if payload:
            status = normalize(payload.get("status"))
            if status:
                note_parts.append(f"adapter_status={status}")
        if not payload and stdout:
            note_parts.append(stdout[:500])
        if stderr:
            note_parts.append("STDERR: " + stderr[:500])

        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result=result,
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes=" | ".join(note_parts) or f"Adapter exit code {completed.returncode}",
        )
    except subprocess.TimeoutExpired:
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Fail",
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes="Timed out.",
        )
    except (subprocess.SubprocessError, OSError, ValueError, TypeError) as exc:
        return TestResult(
            test_id=test.test_id,
            platform=test.platform,
            priority=test.priority,
            automation=test.automation,
            result="Fail",
            runtime_seconds=round(time.perf_counter() - start, 3),
            notes=f"Runner exception: {exc}",
        )

def write_results_csv(results: List[TestResult], path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["test_id","platform","priority","automation","result","runtime_seconds","notes"])
        for r in results:
            w.writerow([r.test_id, r.platform, r.priority, r.automation, r.result, r.runtime_seconds, r.notes])

def write_results_json(results: List[TestResult], path: str):
    summary = {
        "total": len(results),
        "passed": sum(1 for r in results if r.result == "Pass"),
        "failed": sum(1 for r in results if r.result == "Fail"),
        "blocked": sum(1 for r in results if r.result == "Blocked"),
        "p0_failed": sum(1 for r in results if r.priority == "P0" and r.result == "Fail"),
        "results": [asdict(r) for r in results],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

def update_workbook(source_path: str, output_path: str, results: List[TestResult]):
    result_map = {r.test_id: r for r in results}
    wb = load_workbook(source_path)
    # Update platform sheets AND "All Tests" so the dashboard stays in sync
    update_sheets = list(DEFAULT_SHEETS)
    if "All Tests" in wb.sheetnames and "All Tests" not in update_sheets:
        update_sheets.append("All Tests")
    for sheet_name in update_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            test_id = normalize(ws.cell(r, 1).value)
            if test_id in result_map:
                res = result_map[test_id]
                ws.cell(r, 12).value = res.result
                ws.cell(r, 13).value = res.notes
    wb.save(output_path)

def print_summary(results: List[TestResult]):
    total = len(results)
    passed = sum(1 for r in results if r.result == "Pass")
    failed = sum(1 for r in results if r.result == "Fail")
    blocked = sum(1 for r in results if r.result == "Blocked")
    p0_failed = sum(1 for r in results if r.priority == "P0" and r.result == "Fail")
    print(json.dumps({
        "total": total,
        "passed": passed,
        "failed": failed,
        "blocked": blocked,
        "p0_failed": p0_failed,
    }, indent=2))

def main():
    ap = argparse.ArgumentParser(description="Run PDF Vector Importer QA tests from workbook.")
    ap.add_argument("--workbook", help="Path to QA workbook .xlsx")
    ap.add_argument(
        "--init-workbook",
        help=(
            "Create a starter QA workbook (.xlsx) at this path and exit. "
            "Useful on a clean clone before first test run."
        ),
    )
    ap.add_argument("--config", help="Path to JSON config for adapters")
    ap.add_argument("--platform", choices=["SU", "FC", "BL", "LC"])
    ap.add_argument("--priority", choices=["P0","P1","P2"])
    ap.add_argument("--automation", choices=["AUTO","MANUAL","HYBRID"])
    ap.add_argument("--dry-run", action="store_true", help="Validate selection and adapters without executing commands")
    ap.add_argument("--validate-only", action="store_true", help="Only validate workbook schema and exit")
    ap.add_argument("--results-csv", default="qa_results.csv")
    ap.add_argument("--results-json", default="qa_results.json")
    ap.add_argument("--updated-workbook", help="Write a copy of the workbook with Status/Notes filled from results")
    args = ap.parse_args()

    if args.init_workbook:
        init_workbook(args.init_workbook)
        print(f"Created starter workbook: {Path(args.init_workbook).resolve()}")
        return 0

    if not args.workbook:
        ap.error("--workbook is required unless --init-workbook is used.")

    config = load_config(args.config)
    tests = iter_tests(args.workbook, DEFAULT_SHEETS)

    if args.validate_only:
        print(f"Workbook schema valid. Loaded {len(tests)} tests.")
        return 0

    selected = select_tests(tests, args.platform, args.priority, args.automation)
    if not selected:
        print("No tests selected.", file=sys.stderr)
        return 2

    results = [run_one(t, config, args.dry_run) for t in selected]
    write_results_csv(results, args.results_csv)
    write_results_json(results, args.results_json)
    if args.updated_workbook:
        update_workbook(args.workbook, args.updated_workbook, results)
    print_summary(results)
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
